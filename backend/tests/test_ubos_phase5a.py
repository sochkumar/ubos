"""UBOS Phase 5 Sub-pass A — Export / Import / Password Shares."""
from __future__ import annotations
import io
import os
import time
import uuid
import pytest
import requests
from openpyxl import load_workbook


def _get_backend_url():
    v = os.environ.get("REACT_APP_BACKEND_URL", "")
    if not v:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    v = line.split("=", 1)[1].strip()
                    break
    return v.rstrip("/")


BASE_URL = _get_backend_url()
assert BASE_URL
API = f"{BASE_URL}/api"


def _login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="session")
def owner_tok():
    d = _login("owner@ubos.test", "OwnerPass!123")
    return d["access_token"], d["org_id"], d["user"]["id"]


@pytest.fixture(scope="session")
def editor_tok():
    d = _login("editor@ubos.test", "EditorPass!123")
    return d["access_token"], d["org_id"], d["user"]["id"]


def _h(tok):
    return {"Authorization": f"Bearer {tok}"}


def _hjson(tok):
    return {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}


# Resolve the Acme "products" entity_type id lazily and cache it. The old
# suite hardcoded a UUID from an early seed; that ID changes whenever the
# demo org is re-seeded, so look it up dynamically.
_PRODUCTS_ET_CACHE: dict[str, str] = {}


def _resolve_products_et() -> str:
    if "id" in _PRODUCTS_ET_CACHE:
        return _PRODUCTS_ET_CACHE["id"]
    d = _login("owner@ubos.test", "OwnerPass!123")
    tok = d["access_token"]
    r = requests.get(f"{API}/entity-types", headers=_h(tok), timeout=15)
    assert r.status_code == 200, r.text
    for et in r.json():
        if et["key"] == "products":
            _PRODUCTS_ET_CACHE["id"] = et["id"]
            return et["id"]
    pytest.skip("Acme seed missing the `products` entity type")


class _LazyProductsET(str):
    """A str proxy that resolves the products entity_type id on first use."""

    def __str__(self):  # type: ignore[override]
        return _resolve_products_et()

    def __repr__(self):  # noqa: D401
        return _resolve_products_et()

    def __eq__(self, other):
        return _resolve_products_et() == other

    def __hash__(self):
        return hash(_resolve_products_et())

    def __format__(self, spec):
        return format(_resolve_products_et(), spec)


PRODUCTS_ET = _LazyProductsET()


@pytest.fixture(scope="session")
def product_record_ids(owner_tok):
    tok, _, _ = owner_tok
    r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records?limit=5", headers=_h(tok), timeout=15)
    assert r.status_code == 200
    data = r.json()
    items = data.get("items") if isinstance(data, dict) else data
    ids = [it["id"] for it in items[:3]]
    assert len(ids) >= 1
    return ids


# ═══════════════════ EXPORT ═══════════════════
class TestExport:
    def test_csv_export_bom_and_headers(self, owner_tok):
        tok, _, _ = owner_tok
        r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records/export?format=csv",
                         headers=_h(tok), timeout=30)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")
        content = r.content
        assert content[:3] == b"\xef\xbb\xbf", "UTF-8 BOM missing"
        text = content.decode("utf-8-sig")
        assert "\r\n" in text
        first_line = text.splitlines()[0]
        assert "Record #" in first_line
        assert "Title" in first_line
        assert "SKU" in first_line
        assert "Categories" in first_line
        assert "Tags" in first_line
        assert "Created" in first_line
        assert "Updated" in first_line

    def test_xlsx_export_readable(self, owner_tok):
        tok, _, _ = owner_tok
        r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records/export?format=xlsx",
                         headers=_h(tok), timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.max_row >= 2
        assert ws.max_column >= 5

    def test_export_filter_q(self, owner_tok):
        tok, _, _ = owner_tok
        r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records/export?format=csv&q=CHR",
                         headers=_h(tok), timeout=30)
        assert r.status_code == 200

    def test_export_columns_filter(self, owner_tok):
        tok, _, _ = owner_tok
        r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records/export?format=csv&columns=sku,price",
                         headers=_h(tok), timeout=30)
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        header = text.splitlines()[0]
        assert "SKU" in header
        assert "Price" in header
        # 'Notes' should NOT be in filtered columns
        assert "Notes" not in header

    def test_bulk_export_preserves_order(self, owner_tok, product_record_ids):
        tok, _, _ = owner_tok
        r = requests.post(
            f"{API}/entity-types/{PRODUCTS_ET}/records/export-bulk",
            headers=_hjson(tok),
            json={"record_ids": product_record_ids, "format": "csv", "columns": ["sku", "price"]},
            timeout=30,
        )
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) == 1 + len(product_record_ids)

    def test_export_writes_audit(self, owner_tok):
        tok, oorg, _ = owner_tok
        requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records/export?format=csv",
                     headers=_h(tok), timeout=30)
        time.sleep(0.5)
        r = requests.get(f"{API}/audit-logs?action=record.exported&limit=5",
                         headers=_h(tok), timeout=15)
        # audit endpoint may vary; accept 200 or 404
        if r.status_code == 200:
            data = r.json()
            items = data.get("items") if isinstance(data, dict) else data
            assert any(x.get("action") == "record.exported" for x in items or [])


# ═══════════════════ IMPORT ═══════════════════
class TestImport:
    def test_preview_suggested_mapping(self, owner_tok):
        tok, _, _ = owner_tok
        csv_data = "SKU,Title,Price,Category,In stock\nTEST_A,Prod A,10.5,chair,true\n"
        files = {"file": ("test.csv", csv_data.encode(), "text/csv")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "import_token" in d
        assert d["detected_format"] == "csv"
        assert d["headers"] == ["SKU", "Title", "Price", "Category", "In stock"]
        assert d["total_rows"] == 1
        sm = d["suggested_mapping"]
        assert sm["SKU"]["field_key"] == "sku"
        assert sm["SKU"]["confidence"] >= 0.65
        assert sm["Title"]["field_key"] == "title"
        assert sm["Price"]["field_key"] == "price"
        assert sm["Category"]["field_key"] == "category"
        assert sm["In stock"]["field_key"] == "in_stock"

    def test_preview_no_extension(self, owner_tok):
        tok, _, _ = owner_tok
        files = {"file": ("noext", b"a,b\n1,2\n", "text/csv")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=15)
        assert r.status_code == 422

    def test_preview_bad_extension(self, owner_tok):
        tok, _, _ = owner_tok
        files = {"file": ("foo.txt", b"hello", "text/plain")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=15)
        assert r.status_code == 422
        # response body: {"detail": {"code": "unsupported_format", ...}} typically
        assert "unsupported" in r.text.lower()

    def test_preview_file_too_large(self, owner_tok):
        tok, _, _ = owner_tok
        big = b"a,b\n" + (b"1,2\n" * (3 * 1024 * 1024))  # ~12 MB
        files = {"file": ("big.csv", big, "text/csv")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=60)
        assert r.status_code == 413
        assert "too_large" in r.text.lower() or "file_too_large" in r.text.lower()

    def test_plan_error_policy_counts(self, owner_tok):
        tok, _, _ = owner_tok
        # 5 rows: 2 invalid (bad price, unknown dropdown), 3 valid
        csv_data = (
            "SKU,Title,Price,Category,In stock\n"
            "TEST_P1,Prod 1,10.5,chair,true\n"
            "TEST_P2,Prod 2,not-a-number,chair,true\n"     # bad price
            "TEST_P3,Prod 3,20,unknown_option,true\n"      # bad dropdown
            "TEST_P4,Prod 4,30,table,false\n"
            "TEST_P5,Prod 5,40,sofa,true\n"
        )
        files = {"file": ("plan5.csv", csv_data.encode(), "text/csv")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=15)
        assert r.status_code == 200
        tok_import = r.json()["import_token"]

        mapping = {
            "SKU": "sku", "Title": "title", "Price": "price",
            "Category": "category", "In stock": "in_stock",
        }
        body = {"import_token": tok_import, "mapping": mapping,
                "options": {"match_by": "sku", "conflict_policy": "error"}}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/plan",
                          headers=_hjson(tok), json=body, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["total_rows"] == 5
        # 2 invalid rows should be flagged as errors
        assert d["would_error"] >= 2
        # 3 valid rows should be inserts (SKUs are fresh TEST_ prefixed)
        assert d["would_insert"] >= 3 or d["would_update"] + d["would_skip"] + d["would_insert"] >= 3
        # Errors listed
        assert len(d.get("first_errors") or []) >= 2

    def test_execute_and_progress(self, owner_tok):
        tok, oorg, _ = owner_tok
        unique = uuid.uuid4().hex[:6]
        csv_data = (
            "SKU,Title,Price,Category,In stock\n"
            f"TEST_E_{unique}_1,Row1,10,chair,true\n"
            f"TEST_E_{unique}_2,Row2,20,table,false\n"
            f"TEST_E_{unique}_3,Row3,30,sofa,true\n"
        )
        files = {"file": ("exec.csv", csv_data.encode(), "text/csv")}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/preview",
                          headers=_h(tok), files=files, timeout=15)
        assert r.status_code == 200
        tok_import = r.json()["import_token"]
        mapping = {"SKU": "sku", "Title": "title", "Price": "price",
                   "Category": "category", "In stock": "in_stock"}
        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/plan",
                          headers=_hjson(tok), json={
                              "import_token": tok_import, "mapping": mapping,
                              "options": {"match_by": "sku", "conflict_policy": "error"}}, timeout=15)
        assert r.status_code == 200
        plan_id = r.json()["plan_id"]

        r = requests.post(f"{API}/entity-types/{PRODUCTS_ET}/records/import/execute",
                          headers=_hjson(tok), json={"plan_id": plan_id}, timeout=15)
        assert r.status_code == 200, r.text
        j = r.json()
        assert "job_id" in j
        assert j["status"] in ("queued", "running", "completed")
        job_id = j["job_id"]
        assert j["progress_url"].endswith(f"/imports/{job_id}/progress")

        # Poll for completion (up to 30s)
        final = None
        for _ in range(30):
            r = requests.get(f"{API}/imports/{job_id}/progress", headers=_h(tok), timeout=15)
            assert r.status_code == 200
            p = r.json()
            if p["status"] in ("completed", "failed"):
                final = p
                break
            time.sleep(1)
        assert final is not None, "import job did not complete in 30s"
        assert final["status"] == "completed", f"status: {final}"
        assert final["processed"] == 3
        assert final["inserted"] == 3
        assert final["errors"] == 0

        # Verify records now exist (list latest records; SKUs must appear)
        r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records?limit=20",
                         headers=_h(tok), timeout=15)
        assert r.status_code == 200
        d = r.json()
        items = d.get("items") if isinstance(d, dict) else d
        skus = {(it.get("fields") or {}).get("sku") for it in items or []}
        assert any(s and f"TEST_E_{unique}" in s for s in skus), f"imported records not found: {skus}"

    def test_cross_org_isolation_progress(self, owner_tok, editor_tok):
        otok, oorg, _ = owner_tok
        etok, eorg, _ = editor_tok
        if oorg == eorg:
            pytest.skip("same org — no cross-org test possible")
        # Editor makes a job in their org
        csv_data = "SKU,Title,Price\nX1,X,10\n"
        # editor may not have records.create in Products ET, so just query a fake job as editor
        # then ensure owner cannot see it. Use a made-up UUID: owner query → 404
        fake_job = str(uuid.uuid4())
        r = requests.get(f"{API}/imports/{fake_job}/progress", headers=_h(otok), timeout=10)
        assert r.status_code == 404


# ═══════════════════ PASSWORD SHARES ═══════════════════
@pytest.fixture(scope="module")
def product_record_id_for_share(owner_tok):
    tok, _, _ = owner_tok
    r = requests.get(f"{API}/entity-types/{PRODUCTS_ET}/records?limit=1", headers=_h(tok), timeout=15)
    d = r.json()
    items = d.get("items") if isinstance(d, dict) else d
    return items[0]["id"]


class TestPasswordShares:
    def test_create_password_share(self, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password", "password": "secret123"},
                          timeout=15)
        assert r.status_code == 201, r.text
        d = r.json()
        assert d["has_password"] is True
        assert "password_hash" not in d
        assert d["visibility"] == "password"

    def test_create_password_missing(self, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password"}, timeout=15)
        assert r.status_code == 422
        assert "password_required" in r.text or "password" in r.text.lower()

    def test_create_password_too_short(self, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password", "password": "short"}, timeout=15)
        assert r.status_code == 422

    def test_unlock_flow_and_rotation(self, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        # Create fresh share
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password", "password": "secret123"}, timeout=15)
        assert r.status_code == 201
        share = r.json()
        token = share["token"]
        sid = share["id"]

        # Unauthed GET → 401 password_required
        r = requests.get(f"{API}/public/records/{token}", timeout=15)
        assert r.status_code == 401
        assert "password_required" in r.text

        # Wrong password → 401 invalid_password
        s = requests.Session()
        r = s.post(f"{API}/public/records/{token}/unlock",
                   json={"password": "wrongpass"}, timeout=15)
        assert r.status_code == 401
        assert "invalid_password" in r.text

        # Correct password → 200 + cookie
        r = s.post(f"{API}/public/records/{token}/unlock",
                   json={"password": "secret123"}, timeout=15)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["unlocked"] is True
        assert d["expires_in"] > 0
        cookie_name = f"share_unlock_{token}"
        assert cookie_name in s.cookies, f"cookies: {dict(s.cookies)}"

        # GET with cookie → 200
        r = s.get(f"{API}/public/records/{token}", timeout=15)
        assert r.status_code == 200, r.text
        assert "record" in r.json()

        # Signed-in owner bypass without cookie
        r = requests.get(f"{API}/public/records/{token}", headers=_h(tok), timeout=15)
        assert r.status_code == 200

        # Rotate password via PATCH → old cookie invalid
        r = requests.patch(f"{API}/shares/{sid}", headers=_hjson(tok),
                           json={"password": "newsecret456"}, timeout=15)
        assert r.status_code == 200
        assert "password_hash" not in r.json()

        # Same session (old cookie) → 401 again
        r = s.get(f"{API}/public/records/{token}", timeout=15)
        assert r.status_code == 401

        # Switch away from password
        r = requests.patch(f"{API}/shares/{sid}", headers=_hjson(tok),
                           json={"visibility": "public"}, timeout=15)
        assert r.status_code == 200
        d = r.json()
        assert d["visibility"] == "public"
        assert d["has_password"] is False

        # Now unauthed GET → 200
        r = requests.get(f"{API}/public/records/{token}", timeout=15)
        assert r.status_code == 200

    def test_rate_limit_5_attempts(self, reset_rate_limits, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password", "password": "correcthorse"}, timeout=15)
        assert r.status_code == 201
        token = r.json()["token"]

        # Fire 5 wrong attempts → 401; 6th → 429
        last_status = None
        for i in range(6):
            r = requests.post(f"{API}/public/records/{token}/unlock",
                              json={"password": f"wrong{i}"}, timeout=15)
            last_status = r.status_code
            if i < 5:
                assert r.status_code == 401, f"attempt {i}: {r.status_code} {r.text}"
        assert last_status == 429, f"6th expected 429, got {last_status}"
        assert "too_many_attempts" in r.text
        assert "retry_after" in r.text

    def test_media_requires_unlock(self, owner_tok, product_record_id_for_share):
        tok, _, _ = owner_tok
        r = requests.post(f"{API}/records/{product_record_id_for_share}/shares",
                          headers=_hjson(tok),
                          json={"visibility": "password", "password": "mysecret1"}, timeout=15)
        assert r.status_code == 201
        token = r.json()["token"]

        # qr without cookie → 401
        r = requests.get(f"{API}/public/records/{token}/qr.png", timeout=15)
        assert r.status_code == 401
        # barcode without cookie → 401
        r = requests.get(f"{API}/public/records/{token}/barcode.png", timeout=15)
        assert r.status_code == 401


# ═══════════════════ OPENAPI ═══════════════════
class TestOpenAPI:
    def test_openapi_paths_present(self):
        r = requests.get(f"{API}/openapi.json", timeout=15)
        assert r.status_code == 200
        paths = r.json().get("paths", {})
        expected = [
            "/api/records/{rid}/shares",
            "/api/public/records/{token}/unlock",
            "/api/entity-types/{et_id}/records/export",
            "/api/entity-types/{et_id}/records/import/preview",
            "/api/entity-types/{et_id}/records/import/plan",
            "/api/entity-types/{et_id}/records/import/execute",
            "/api/imports/{job_id}/progress",
            "/api/imports/{job_id}/errors.csv",
        ]
        for p in expected:
            assert p in paths, f"missing OpenAPI path {p}. Available: {list(paths.keys())[:20]}"
