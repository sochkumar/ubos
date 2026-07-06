"""CSV / Excel import + export for records — Phase 5 Sub-pass A.

Exports stream via StreamingResponse; imports use a 3-step wizard
(preview → plan → execute) with progress polling on a background task.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from fastapi import (
    APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query, Request,
    UploadFile, File,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from rapidfuzz import fuzz, process

from audit import audit
from auth_deps import AuthContext, require_permission
from core.storage.factory import get_storage_adapter
from db import get_db, tenant_filter
from models import strip_id
from services.query_builder import build_filter_query, build_sort_spec
from services.categories import descendant_ids_including_self
from routes.data import _build_records_query, _load_field_defs

router = APIRouter(tags=["export-import"])


# ═══════════════════════════════════════════════════════════════════════
#  Configuration
# ═══════════════════════════════════════════════════════════════════════
EXPORT_MAX_ROWS = int(os.environ.get("EXPORT_MAX_ROWS", "50000"))
IMPORT_MAX_FILE_MB = int(os.environ.get("IMPORT_MAX_FILE_MB", "10"))
IMPORT_MAX_ROWS = int(os.environ.get("IMPORT_MAX_ROWS", "50000"))
IMPORT_TMP_ROOT = Path(os.environ.get("IMPORT_TMP_ROOT", "/tmp/ubos_imports"))
IMPORT_TOKEN_TTL_SEC = 30 * 60  # 30 min


# ═══════════════════════════════════════════════════════════════════════
#  Value formatting for export
# ═══════════════════════════════════════════════════════════════════════

def _fmt_value_for_export(
    fd: dict | None, value: Any,
    media_lookup: dict[str, dict] | None = None,
    signed_url_of: dict[str, str] | None = None,
) -> Any:
    if value is None or value == "":
        return None
    ftype = (fd or {}).get("type", "text")
    if ftype == "boolean":
        return "true" if value else "false"
    if ftype in ("date", "datetime"):
        return value  # ISO already
    if ftype == "currency":
        try:
            return float(value)
        except Exception:
            return value
    if ftype == "number":
        try:
            return float(value) if "." in str(value) else int(value)
        except Exception:
            return value
    if ftype == "dropdown":
        return value
    if ftype == "multi_select":
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)
    if ftype in ("image", "file"):
        # value expected to be a media_id or list of media_ids
        ids = value if isinstance(value, list) else [value]
        parts: list[str] = []
        for mid in ids:
            m = (media_lookup or {}).get(mid)
            if not m:
                parts.append(mid)
                continue
            url = (signed_url_of or {}).get(mid, "")
            parts.append(f"{m.get('filename')} <{url}>" if url else m.get("filename", mid))
        return "; ".join(parts)
    if ftype == "relation":
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)
    return str(value)


def _column_headers(field_defs: list[dict], include_metadata: bool) -> tuple[list[str], list[str]]:
    """Return (keys, human_labels)."""
    keys = ["record_number", "title"] if include_metadata else []
    labels = ["Record #", "Title"] if include_metadata else []
    for fd in field_defs:
        keys.append(fd["key"])
        labels.append(fd.get("label") or fd["key"])
    if include_metadata:
        keys.extend(["categories", "tags", "created_at", "updated_at"])
        labels.extend(["Categories", "Tags", "Created", "Updated"])
    return keys, labels


async def _resolve_export_ctx(
    db, org_id: str, et_id: str, records: list[dict], include_media_urls: bool,
) -> tuple[dict[str, dict], dict[str, str], dict[str, dict], dict[str, dict]]:
    """Prefetch media, signed URLs, categories, and tags used by the export."""
    media_ids: set[str] = set()
    cat_ids: set[str] = set()
    tag_ids: set[str] = set()
    for r in records:
        for v in (r.get("fields") or {}).values():
            if isinstance(v, str) and len(v) == 36 and v.count("-") == 4:
                media_ids.add(v)
            elif isinstance(v, list):
                for item in v:
                    if isinstance(item, str) and len(item) == 36 and item.count("-") == 4:
                        media_ids.add(item)
        cat_ids.update(r.get("category_ids") or [])
        tag_ids.update(r.get("tag_ids") or [])
    media_lookup: dict[str, dict] = {}
    signed_urls: dict[str, str] = {}
    if media_ids and include_media_urls:
        docs = await db.media.find(
            {"_id": {"$in": list(media_ids)}, "org_id": org_id},
            {"filename": 1, "mime": 1, "storage_key": 1, "size": 1},
        ).to_list(len(media_ids))
        adapter = get_storage_adapter()
        for m in docs:
            media_lookup[m["_id"]] = m
            try:
                signed_urls[m["_id"]] = await adapter.presigned_get(m["storage_key"], ttl_seconds=1800)
            except Exception:
                signed_urls[m["_id"]] = ""
    cats: dict[str, dict] = {}
    if cat_ids:
        for c in await db.categories.find(
            {"_id": {"$in": list(cat_ids)}, "org_id": org_id},
            {"name": 1, "path_names": 1},
        ).to_list(len(cat_ids)):
            cats[c["_id"]] = c
    tags: dict[str, dict] = {}
    if tag_ids:
        for t in await db.tags.find(
            {"_id": {"$in": list(tag_ids)}, "org_id": org_id},
            {"name": 1},
        ).to_list(len(tag_ids)):
            tags[t["_id"]] = t
    return media_lookup, signed_urls, cats, tags


def _row_for_record(
    rec: dict, keys: list[str], defs_by_key: dict[str, dict],
    include_metadata: bool,
    media_lookup: dict[str, dict], signed_urls: dict[str, str],
    cats: dict[str, dict], tags: dict[str, dict],
) -> list[Any]:
    row: list[Any] = []
    fields = rec.get("fields") or {}
    for k in keys:
        if k == "record_number":
            row.append(rec.get("record_number"))
        elif k == "title":
            row.append(rec.get("title"))
        elif k == "categories":
            cat_labels = []
            for cid in (rec.get("category_ids") or []):
                c = cats.get(cid)
                if c:
                    if c.get("path_names"):
                        cat_labels.append(" / ".join(c["path_names"]))
                    else:
                        cat_labels.append(c.get("name", cid))
            row.append("; ".join(cat_labels))
        elif k == "tags":
            row.append(", ".join(tags.get(t, {}).get("name", t) for t in (rec.get("tag_ids") or [])))
        elif k == "created_at":
            row.append(rec.get("created_at"))
        elif k == "updated_at":
            row.append(rec.get("updated_at"))
        else:
            fd = defs_by_key.get(k)
            row.append(_fmt_value_for_export(fd, fields.get(k), media_lookup, signed_urls))
    return row


# ═══════════════════════════════════════════════════════════════════════
#  Export streamers
# ═══════════════════════════════════════════════════════════════════════

def _csv_stream(headers: list[str], rows_gen: Iterator[list[Any]]) -> Iterator[bytes]:
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    # UTF-8 BOM so Excel opens the file with proper encoding
    yield "\ufeff".encode("utf-8")
    writer.writerow(headers)
    yield buf.getvalue().encode("utf-8")
    buf.seek(0)
    buf.truncate()
    for row in rows_gen:
        writer.writerow([("" if v is None else v) for v in row])
        yield buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate()


def _xlsx_bytes(headers: list[str], defs_by_key: dict[str, dict], key_order: list[str],
                rows_gen: Iterator[list[Any]]) -> bytes:
    """Build the workbook in write-only mode and return bytes."""
    from openpyxl.cell import WriteOnlyCell
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Records")
    header_row = [WriteOnlyCell(ws, value=h) for h in headers]
    for c in header_row:
        c.font = c.font.copy(bold=True)
    ws.append(header_row)

    date_cols = {
        i for i, k in enumerate(key_order)
        if defs_by_key.get(k, {}).get("type") in ("date", "datetime")
        or k in ("created_at", "updated_at")
    }
    num_cols = {
        i for i, k in enumerate(key_order)
        if defs_by_key.get(k, {}).get("type") in ("number", "currency")
    }

    # Track max widths for rough autofit
    widths = [max(len(h), 8) for h in headers]

    for row in rows_gen:
        out_row = []
        for i, v in enumerate(row):
            if v is None or v == "":
                out_row.append(None)
                continue
            if i in date_cols and isinstance(v, str):
                try:
                    dt = datetime.fromisoformat(v.replace("Z", "+00:00"))
                    if dt.tzinfo is not None:
                        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
                    cell = WriteOnlyCell(ws, value=dt)
                    cell.number_format = "yyyy-mm-dd hh:mm" if "T" in v else "yyyy-mm-dd"
                    out_row.append(cell)
                    widths[i] = max(widths[i], 18)
                    continue
                except Exception:
                    pass
            if i in num_cols:
                try:
                    out_row.append(float(v) if "." in str(v) else int(v))
                    widths[i] = max(widths[i], len(str(v)) + 2)
                    continue
                except Exception:
                    pass
            s = str(v)
            widths[i] = max(widths[i], min(60, len(s) + 2))
            out_row.append(s)
        ws.append(out_row)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w, 8), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════
#  Export endpoints
# ═══════════════════════════════════════════════════════════════════════

@router.get("/entity-types/{et_id}/records/export")
async def export_records(
    et_id: str, request: Request, bg: BackgroundTasks,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    q: str | None = Query(None),
    category_id: str | None = Query(None),
    tag_ids: str | None = Query(None, description="Comma-separated tag IDs"),
    columns: str | None = Query(None, description="Comma-separated field keys to include (default: all)"),
    include_metadata: bool = Query(True),
    limit: int = Query(EXPORT_MAX_ROWS, ge=1, le=EXPORT_MAX_ROWS),
    ctx: AuthContext = Depends(require_permission("records.read")),
):
    db = get_db()
    et = await db.entity_types.find_one(tenant_filter(ctx.org_id, {"_id": et_id}),
                                        {"_id": 1, "key": 1})
    if not et:
        raise HTTPException(404, "entity type not found")

    tags = [t.strip() for t in (tag_ids or "").split(",") if t.strip()] or None
    filt = await _build_records_query(db, ctx, et_id, q, category_id, tags, [])

    all_defs = await _load_field_defs(db, ctx.org_id, et_id)
    if columns:
        want = [c.strip() for c in columns.split(",") if c.strip()]
        defs = [d for d in all_defs if d["key"] in want]
        defs.sort(key=lambda d: want.index(d["key"]))
    else:
        defs = list(all_defs)
    defs_by_key = {d["key"]: d for d in all_defs}
    key_order, labels = _column_headers(defs, include_metadata)

    total = await db.records.count_documents(filt)
    total = min(total, limit)

    records = await db.records.find(filt).sort("created_at", -1).limit(limit).to_list(limit)
    media_lookup, signed_urls, cats, tags_lookup = await _resolve_export_ctx(
        db, ctx.org_id, et_id, records, include_media_urls=True,
    )

    def rows_gen() -> Iterator[list[Any]]:
        for r in records:
            yield _row_for_record(
                r, key_order, defs_by_key, include_metadata,
                media_lookup, signed_urls, cats, tags_lookup,
            )

    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    fname = f"{et.get('key') or 'records'}-export-{ts}.{format}"

    audit(bg, action="record.exported", actor_id=ctx.user["_id"], org_id=ctx.org_id,
          target_type="entity_type", target_id=et_id,
          diff={"count": total, "format": format, "filters": {"q": q, "category_id": category_id, "tag_ids": tags}},
          request=request)

    if format == "csv":
        return StreamingResponse(
            _csv_stream(labels, rows_gen()),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    # xlsx
    body = _xlsx_bytes(labels, defs_by_key, key_order, rows_gen())
    return StreamingResponse(
        iter([body]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


class BulkExportBody(BaseModel):
    record_ids: list[str] = Field(min_length=1, max_length=EXPORT_MAX_ROWS)
    format: str = Field("csv", pattern="^(csv|xlsx)$")
    columns: list[str] | None = None
    include_metadata: bool = True


@router.post("/entity-types/{et_id}/records/export-bulk")
async def export_records_bulk(
    et_id: str, body: BulkExportBody, request: Request, bg: BackgroundTasks,
    ctx: AuthContext = Depends(require_permission("records.read")),
):
    db = get_db()
    et = await db.entity_types.find_one(tenant_filter(ctx.org_id, {"_id": et_id}),
                                        {"_id": 1, "key": 1})
    if not et:
        raise HTTPException(404, "entity type not found")
    filt = tenant_filter(ctx.org_id, {"entity_type_id": et_id, "_id": {"$in": body.record_ids}})
    records = await db.records.find(filt).to_list(len(body.record_ids))
    # Preserve caller ordering
    idx = {rid: i for i, rid in enumerate(body.record_ids)}
    records.sort(key=lambda r: idx.get(r["_id"], 10**9))

    all_defs = await _load_field_defs(db, ctx.org_id, et_id)
    if body.columns:
        defs = [d for d in all_defs if d["key"] in body.columns]
        defs.sort(key=lambda d: body.columns.index(d["key"]))
    else:
        defs = list(all_defs)
    defs_by_key = {d["key"]: d for d in all_defs}
    key_order, labels = _column_headers(defs, body.include_metadata)

    media_lookup, signed_urls, cats, tags_lookup = await _resolve_export_ctx(
        db, ctx.org_id, et_id, records, include_media_urls=True,
    )
    def rows_gen():
        for r in records:
            yield _row_for_record(r, key_order, defs_by_key, body.include_metadata,
                                  media_lookup, signed_urls, cats, tags_lookup)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    fname = f"{et.get('key') or 'records'}-selected-{ts}.{body.format}"

    audit(bg, action="record.exported", actor_id=ctx.user["_id"], org_id=ctx.org_id,
          target_type="entity_type", target_id=et_id,
          diff={"count": len(records), "format": body.format, "kind": "bulk"},
          request=request)

    if body.format == "csv":
        return StreamingResponse(_csv_stream(labels, rows_gen()),
                                 media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    body_bytes = _xlsx_bytes(labels, defs_by_key, key_order, rows_gen())
    return StreamingResponse(
        iter([body_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ═══════════════════════════════════════════════════════════════════════
#  Import: helpers
# ═══════════════════════════════════════════════════════════════════════

def _import_dir() -> Path:
    IMPORT_TMP_ROOT.mkdir(parents=True, exist_ok=True)
    return IMPORT_TMP_ROOT


def _cache_path(token: str, suffix: str) -> Path:
    return _import_dir() / f"{token}{suffix}"


def _write_cache_meta(token: str, meta: dict) -> None:
    _cache_path(token, ".meta.json").write_text(json.dumps(meta))


def _read_cache_meta(token: str) -> dict:
    p = _cache_path(token, ".meta.json")
    if not p.exists():
        raise HTTPException(404, "import token expired or invalid")
    meta = json.loads(p.read_text())
    if meta.get("expires_at", 0) < time.time():
        raise HTTPException(410, "import token expired")
    return meta


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _suggest_mapping(headers: list[str], field_defs: list[dict]) -> dict[str, dict]:
    """Fuzzy match each CSV header to a field key/label. Confidence 0-1."""
    if not headers:
        return {}
    choices: dict[str, str] = {
        # Special record-level fields
        "title": "title", "name": "title",
        "record number": "record_number", "record_number": "record_number", "record #": "record_number",
        "tags": "tags",
    }
    for fd in field_defs or []:
        for alias in filter(None, [fd.get("label"), fd["key"], _slug(fd.get("label")), _slug(fd["key"])]):
            choices.setdefault(alias, fd["key"])
    picks: dict[str, dict] = {}
    for h in headers:
        h_norm = h.strip()
        best = process.extractOne(h_norm, list(choices.keys()), scorer=fuzz.WRatio)
        if best:
            _, score, _ = best
            picks[h] = {
                "field_key": choices[best[0]] if score >= 65 else None,
                "confidence": round(score / 100.0, 2),
                "reason": f"matched '{best[0]}' ({score}%)" if score >= 65 else "low confidence",
            }
        else:
            picks[h] = {"field_key": None, "confidence": 0.0, "reason": "no match"}
    return picks


def _parse_csv_to_rows(path: Path) -> tuple[list[str], list[dict], int]:
    """Read all rows into memory (bounded by IMPORT_MAX_ROWS)."""
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return [], [], 0
        headers = [h.strip() for h in headers]
        for i, r in enumerate(reader):
            if i >= IMPORT_MAX_ROWS:
                break
            row = {headers[j] if j < len(headers) else f"_col{j}": (v.strip() if isinstance(v, str) else v)
                   for j, v in enumerate(r)}
            rows.append(row)
    return headers, rows, len(rows)


def _parse_xlsx_to_rows(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict], int, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb[sheet_names[0]]
    it = ws.iter_rows(values_only=True)
    try:
        headers_row = next(it)
    except StopIteration:
        wb.close()
        return [], [], 0, sheet_names
    headers = [(str(h).strip() if h is not None else f"_col{i}") for i, h in enumerate(headers_row)]
    rows: list[dict] = []
    for i, r in enumerate(it):
        if i >= IMPORT_MAX_ROWS:
            break
        row = {}
        for j, v in enumerate(r):
            key = headers[j] if j < len(headers) else f"_col{j}"
            if isinstance(v, datetime):
                v = v.isoformat()
            row[key] = v if v is not None else ""
        rows.append(row)
    wb.close()
    return headers, rows, len(rows), sheet_names


# ═══════════════════════════════════════════════════════════════════════
#  Import endpoints
# ═══════════════════════════════════════════════════════════════════════

@router.post("/entity-types/{et_id}/records/import/preview")
async def import_preview(
    et_id: str,
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(require_permission("records.create")),
):
    db = get_db()
    if not await db.entity_types.find_one(tenant_filter(ctx.org_id, {"_id": et_id}), {"_id": 1}):
        raise HTTPException(404, "entity type not found")

    filename = file.filename or "upload"
    if "." not in filename:
        raise HTTPException(422, "filename must have an extension")
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx"):
        raise HTTPException(422, {"code": "unsupported_format",
                                   "detail": "Only CSV or XLSX files are supported."})

    data = await file.read()
    if len(data) > IMPORT_MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(413, {"code": "file_too_large",
                                   "detail": f"Max upload size is {IMPORT_MAX_FILE_MB} MB."})

    token = str(uuid.uuid4())
    cached = _cache_path(token, f".{ext}")
    cached.write_bytes(data)

    sheet_names: list[str] = []
    selected_sheet: str | None = None
    if ext == "csv":
        headers, rows, total = _parse_csv_to_rows(cached)
    else:
        headers, rows, total, sheet_names = _parse_xlsx_to_rows(cached)
        selected_sheet = sheet_names[0] if sheet_names else None

    field_defs = await _load_field_defs(db, ctx.org_id, et_id)
    suggested = _suggest_mapping(headers, field_defs)
    warnings: list[str] = []
    if total >= IMPORT_MAX_ROWS:
        warnings.append(f"Row count truncated to {IMPORT_MAX_ROWS} — split large files or contact support.")

    meta = {
        "token": token,
        "org_id": ctx.org_id,
        "entity_type_id": et_id,
        "user_id": ctx.user["_id"],
        "format": ext,
        "filename": filename,
        "sheet_names": sheet_names,
        "selected_sheet": selected_sheet,
        "expires_at": time.time() + IMPORT_TOKEN_TTL_SEC,
    }
    _write_cache_meta(token, meta)

    resp: dict[str, Any] = {
        "import_token": token,
        "detected_format": ext,
        "headers": headers,
        "preview_rows": rows[:5],
        "total_rows": total,
        "suggested_mapping": suggested,
        "warnings": warnings,
    }
    if sheet_names:
        resp["sheet_names"] = sheet_names
        resp["selected_sheet"] = selected_sheet
    return resp


class ImportPlanBody(BaseModel):
    import_token: str
    mapping: dict[str, str | None]  # csv_header → field_key (or None to ignore)
    options: dict[str, Any] = Field(default_factory=dict)


@router.post("/entity-types/{et_id}/records/import/plan")
async def import_plan(et_id: str, body: ImportPlanBody,
                     ctx: AuthContext = Depends(require_permission("records.create"))):
    db = get_db()
    meta = _read_cache_meta(body.import_token)
    if meta["org_id"] != ctx.org_id or meta["entity_type_id"] != et_id:
        raise HTTPException(403, "import token doesn't belong to this context")

    opts = body.options or {}
    match_by: str | None = opts.get("match_by") or None
    conflict_policy = (opts.get("conflict_policy") or "error").lower()
    auto_create_tags = bool(opts.get("auto_create_tags", True))
    default_values: dict[str, Any] = opts.get("default_values") or {}
    sheet_name = opts.get("sheet_name")

    if conflict_policy not in ("skip", "update", "error"):
        raise HTTPException(422, "conflict_policy must be skip|update|error")

    ext = meta["format"]
    cached = _cache_path(body.import_token, f".{ext}")
    if ext == "csv":
        headers, rows, total = _parse_csv_to_rows(cached)
    else:
        headers, rows, total, _ = _parse_xlsx_to_rows(cached, sheet_name)
        if sheet_name:
            meta["selected_sheet"] = sheet_name
            _write_cache_meta(body.import_token, meta)

    field_defs = await _load_field_defs(db, ctx.org_id, et_id)
    defs_by_key = {d["key"]: d for d in field_defs}
    # Media/relation fields are not importable in this pass
    unsupported = {d["key"] for d in field_defs if d["type"] in ("image", "file", "relation")}

    required_keys = {d["key"] for d in field_defs if d.get("required")}
    mapped_field_keys = {v for v in body.mapping.values() if v}
    missing_required = required_keys - mapped_field_keys - set(default_values.keys())

    warnings: list[str] = []
    for h, k in body.mapping.items():
        if k in unsupported:
            warnings.append(f"Column '{h}' → {k} is a media/relation field; skipped in this pass.")
    if missing_required:
        warnings.append(f"Unmapped required fields: {', '.join(sorted(missing_required))}")

    # Load existing records by match_by index for the plan simulation
    existing_by_key: dict[str, dict] = {}
    if match_by:
        match_def = defs_by_key.get(match_by)
        if not match_def and match_by != "record_number":
            raise HTTPException(422, f"match_by '{match_by}' isn't a defined field or record_number")
        proj_key = f"fields.{match_by}" if match_by != "record_number" else "record_number"
        cursor = db.records.find(
            tenant_filter(ctx.org_id, {"entity_type_id": et_id}),
            {proj_key: 1, "_id": 1},
        )
        async for r in cursor:
            key_val = r.get("record_number") if match_by == "record_number" else (r.get("fields") or {}).get(match_by)
            if key_val:
                existing_by_key[str(key_val)] = r

    would_insert = would_update = would_skip = would_error = 0
    per_row_out: list[dict] = []
    first_errors: list[dict] = []

    # In-batch uniqueness tracking for `unique: true` fields (and match_by
    # if it's a proper unique field). Maps field_key → {value → first_row_idx}.
    unique_field_keys = {d["key"] for d in field_defs if d.get("unique") and d["key"] not in unsupported}
    # If match_by is unique and conflict_policy=update, we treat later
    # duplicates as pointing at the same DB row (all collapse into one update).
    seen_in_batch: dict[str, dict[str, int]] = {k: {} for k in unique_field_keys}

    # Simulate validation for each row
    for idx, csv_row in enumerate(rows):
        payload_fields: dict[str, Any] = {}
        for header, val in csv_row.items():
            field_key = body.mapping.get(header)
            if not field_key or field_key in unsupported:
                continue
            if field_key == "title":
                # Title is captured during execute; the plan just needs to know
                # the row shape and per-field validation.
                continue
            payload_fields[field_key] = _coerce_import_value(defs_by_key.get(field_key), val)
        # Merge default values for empty cells
        for fk, dv in default_values.items():
            if payload_fields.get(fk) in (None, "", []) and fk not in unsupported:
                payload_fields[fk] = dv

        # Validate + emit per-row errors
        errors: list[dict] = []
        for fd in field_defs:
            if fd["key"] in unsupported:
                continue
            v = payload_fields.get(fd["key"])
            if fd.get("required") and (v is None or v == ""):
                errors.append({"field": fd["key"], "msg": "required"})
                continue
            if v is None or v == "":
                continue
            valid, coerced, err = _validate_field_value(fd, v)
            if not valid:
                errors.append({"field": fd["key"], "msg": err or "invalid"})
            else:
                payload_fields[fd["key"]] = coerced

        # In-batch unique-value detection.  For each `unique:true` field
        # the first row's value is recorded; later rows with the same
        # value get an action based on conflict_policy.
        dup_key: str | None = None
        dup_first_row: int | None = None
        for uk in unique_field_keys:
            v = payload_fields.get(uk)
            if v is None or v == "":
                continue
            first = seen_in_batch[uk].get(str(v))
            if first is not None:
                dup_key = uk
                dup_first_row = first
                break
            seen_in_batch[uk][str(v)] = idx

        # Match / conflict simulation
        action = "insert"
        record_id = None
        if match_by:
            key_val = payload_fields.get(match_by) if match_by != "record_number" else csv_row.get(
                next((h for h, k in body.mapping.items() if k == "record_number"), "")
            )
            if key_val and str(key_val) in existing_by_key:
                record_id = existing_by_key[str(key_val)]["_id"]
                if conflict_policy == "skip":
                    action = "skip"
                elif conflict_policy == "update":
                    action = "update"
                else:
                    action = "error"
                    errors.append({"field": match_by, "msg": f"duplicate {match_by}: {key_val}"})

        # Apply in-batch duplicate policy AFTER DB match decision
        if dup_key and not errors:
            if conflict_policy == "skip":
                action = "skip"
            elif conflict_policy == "update" and match_by == dup_key:
                # Later duplicates collapse into the same underlying update.
                # If the first row will be an insert (not matched vs DB),
                # this row is redundant → skip; else this row's update just
                # overwrites the same DB row so we treat it as skip.
                action = "skip"
            else:
                # error, or update-with-different-match-by, or insert path:
                # collide → error out.
                errors.append({
                    "field": dup_key,
                    "msg": f"duplicate unique value '{payload_fields.get(dup_key)}' at row {idx + 1} "
                           f"(already at row {(dup_first_row or 0) + 1} in this file)",
                })
                action = "error"

        if errors:
            action = "error"
            would_error += 1
            if len(first_errors) < 20:
                first_errors.append({"row_idx": idx, "errors": errors})
        elif action == "insert":
            would_insert += 1
        elif action == "update":
            would_update += 1
        elif action == "skip":
            would_skip += 1

        if len(per_row_out) < 100:
            per_row_out.append({
                "row_idx": idx, "action": action,
                "record_id": record_id, "errors": errors or None,
            })

    plan_id = str(uuid.uuid4())
    plan = {
        "plan_id": plan_id,
        "import_token": body.import_token,
        "mapping": body.mapping,
        "options": opts,
        "match_by": match_by,
        "conflict_policy": conflict_policy,
        "auto_create_tags": auto_create_tags,
        "default_values": default_values,
        "sheet_name": sheet_name,
        "totals": {
            "total_rows": total,
            "would_insert": would_insert,
            "would_update": would_update,
            "would_skip": would_skip,
            "would_error": would_error,
        },
        "warnings": warnings,
        "expires_at": time.time() + IMPORT_TOKEN_TTL_SEC,
    }
    _cache_path(plan_id, ".plan.json").write_text(json.dumps(plan))

    return {
        "plan_id": plan_id,
        "total_rows": total,
        "would_insert": would_insert,
        "would_update": would_update,
        "would_skip": would_skip,
        "would_error": would_error,
        "per_row": per_row_out,
        "first_errors": first_errors,
        "warnings": warnings,
    }


class ImportExecuteBody(BaseModel):
    plan_id: str


# ── Import job registry: in-process concurrency guard ───────────────
_IMPORT_JOBS: dict[str, asyncio.Task] = {}


@router.post("/entity-types/{et_id}/records/import/execute")
async def import_execute(et_id: str, body: ImportExecuteBody, request: Request, bg: BackgroundTasks,
                         ctx: AuthContext = Depends(require_permission("records.create"))):
    plan_path = _cache_path(body.plan_id, ".plan.json")
    if not plan_path.exists():
        raise HTTPException(404, "plan_id not found or expired")
    plan = json.loads(plan_path.read_text())
    meta = _read_cache_meta(plan["import_token"])
    if meta["org_id"] != ctx.org_id or meta["entity_type_id"] != et_id:
        raise HTTPException(403, "plan doesn't belong to this context")

    db = get_db()
    # Concurrency guard: max 5 running imports per org
    running = await db.import_jobs.count_documents({
        "org_id": ctx.org_id, "status": {"$in": ["queued", "running"]},
    })
    if running >= 5:
        raise HTTPException(429, {"code": "too_many_running", "detail": "Max 5 concurrent imports."})

    job_id = str(uuid.uuid4())
    job = {
        "_id": job_id,
        "org_id": ctx.org_id,
        "entity_type_id": et_id,
        "user_id": ctx.user["_id"],
        "filename": meta.get("filename"),
        "format": meta.get("format"),
        "mapping": plan.get("mapping"),
        "options": plan.get("options"),
        "status": "queued",
        "total_rows": plan["totals"]["total_rows"],
        "processed": 0, "inserted": 0, "updated": 0,
        "skipped": 0, "errors": 0,
        "error_report_key": None,
        "started_at": None, "completed_at": None, "canceled_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.import_jobs.insert_one(job)

    # Fire the background asyncio task (doesn't block the request)
    task = asyncio.create_task(_run_import_job(job_id, plan, meta))
    _IMPORT_JOBS[job_id] = task

    audit(bg, action="record.imported", actor_id=ctx.user["_id"], org_id=ctx.org_id,
          target_type="entity_type", target_id=et_id,
          diff={"job_id": job_id, "filename": meta.get("filename"),
                "mapping_keys": list(plan.get("mapping", {}).keys()),
                "options": plan.get("options")},
          request=request)

    return {"job_id": job_id, "status": "queued",
            "progress_url": f"/api/imports/{job_id}/progress",
            "started_at": None}


@router.get("/imports/{job_id}/progress")
async def import_progress(job_id: str,
                          ctx: AuthContext = Depends(require_permission("records.read"))):
    db = get_db()
    job = await db.import_jobs.find_one({"_id": job_id, "org_id": ctx.org_id})
    if not job:
        raise HTTPException(404, "job not found")
    out = strip_id(job)
    if job.get("error_report_key"):
        out["error_report_url"] = f"/api/imports/{job_id}/errors.csv"
    return out


@router.get("/imports/{job_id}/errors.csv")
async def import_errors_csv(job_id: str,
                            ctx: AuthContext = Depends(require_permission("records.read"))):
    db = get_db()
    job = await db.import_jobs.find_one({"_id": job_id, "org_id": ctx.org_id})
    if not job or not job.get("error_report_key"):
        raise HTTPException(404, "no error report available")
    p = _cache_path(job["error_report_key"], "")
    if not p.exists():
        raise HTTPException(410, "error report expired")
    return StreamingResponse(
        iter([p.read_bytes()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="import-{job_id}-errors.csv"'},
    )


# ═══════════════════════════════════════════════════════════════════════
#  Import job runner + field validation
# ═══════════════════════════════════════════════════════════════════════

def _coerce_import_value(fd: dict | None, v: Any) -> Any:
    """Coerce raw CSV/XLSX value into the shape expected by FieldValidator."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
    if v == "":
        return None
    ftype = (fd or {}).get("type", "text")
    if ftype == "boolean":
        s = str(v).strip().lower()
        return s in ("true", "1", "yes", "y", "on")
    if ftype in ("number", "currency"):
        try:
            return float(v)
        except Exception:
            return v
    if ftype == "multi_select":
        if isinstance(v, list):
            return v
        return [x.strip() for x in str(v).split(",") if x.strip()]
    return v


def _validate_field_value(fd: dict, v: Any) -> tuple[bool, Any, str | None]:
    """Lightweight per-field validation for the import plan/execute pass.
    Uses `validator.FieldValidator._validate_value` (private but stable enough)
    for type coercion, then falls back on a basic per-type check."""
    ftype = fd.get("type")
    # Dropdowns: always give a friendly, deterministic error.
    if ftype == "dropdown":
        opts = (fd.get("config") or {}).get("options") or []
        if opts and str(v) not in [str(o) for o in opts]:
            return False, None, f"value '{v}' not in dropdown options for field '{fd.get('key')}'"
    try:
        from validator import FieldValidator  # type: ignore
        fv = FieldValidator.__new__(FieldValidator)
        return True, fv._validate_value(fd, v), None
    except ValueError as e:
        return False, None, str(e)
    except Exception:
        # Basic fallback
        if ftype == "number" and not isinstance(v, (int, float)):
            try:
                return True, float(v), None
            except Exception:
                return False, None, "not a number"
        if ftype == "email" and isinstance(v, str) and "@" not in v:
            return False, None, "invalid email"
        return True, v, None


async def _run_import_job(job_id: str, plan: dict, meta: dict) -> None:
    """Background task — actually writes records in batches of 200."""
    db = get_db()
    def now_iso() -> str:
        return datetime.now(timezone.utc).isoformat()
    ctx_org = meta["org_id"]
    et_id = meta["entity_type_id"]
    await db.import_jobs.update_one(
        {"_id": job_id},
        {"$set": {"status": "running", "started_at": now_iso()}},
    )

    try:
        # Re-parse the input to avoid holding memory for hours
        ext = meta["format"]
        cached = _cache_path(plan["import_token"], f".{ext}")
        if ext == "csv":
            headers, rows, _ = _parse_csv_to_rows(cached)
        else:
            headers, rows, _, _ = _parse_xlsx_to_rows(cached, plan.get("sheet_name"))

        field_defs = await _load_field_defs(db, ctx_org, et_id)
        defs_by_key = {d["key"]: d for d in field_defs}
        unsupported = {d["key"] for d in field_defs if d["type"] in ("image", "file", "relation")}

        mapping = plan["mapping"]
        match_by = plan.get("match_by")
        conflict_policy = plan.get("conflict_policy", "error")
        auto_create_tags = plan.get("auto_create_tags", True)
        default_values = plan.get("default_values") or {}

        # Preload existing records by match_by for updates
        existing_by_key: dict[str, dict] = {}
        if match_by:
            proj = "record_number" if match_by == "record_number" else f"fields.{match_by}"
            _projection = {proj: 1, "_id": 1, "record_number": 1}
            if match_by == "record_number":
                _projection["fields"] = 1  # not colliding
            cursor = db.records.find(
                tenant_filter(ctx_org, {"entity_type_id": et_id}),
                _projection,
            )
            async for r in cursor:
                kv = r.get("record_number") if match_by == "record_number" else (r.get("fields") or {}).get(match_by)
                if kv:
                    existing_by_key[str(kv)] = r

        # Tag pre-cache
        tag_map: dict[str, str] = {}  # lowercase name → id
        async for t in db.tags.find({"org_id": ctx_org, "deleted_at": None}, {"name": 1}):
            tag_map[(t.get("name") or "").lower()] = t["_id"]

        error_lines: list[list[str]] = [["row_idx", "field", "message", "raw_value"]]
        inserted = updated = skipped = errors = processed = 0

        # In-batch uniqueness tracking (mirrors the plan phase). Bugfix:
        # without this, two rows with the same value on a `unique:true`
        # field could BOTH end up as inserts.
        unique_field_keys = {d["key"] for d in field_defs
                              if d.get("unique") and d["key"] not in unsupported}
        seen_in_batch: dict[str, dict[str, int]] = {k: {} for k in unique_field_keys}
        # Track record-numbers of updates already scheduled so subsequent
        # collapsing duplicates don't re-schedule the same update.
        collapsed_update_targets: set[str] = set()

        async def _record_number(et_key: str) -> str:
            # Grab a sequential record number — same approach as data.py
            seq = await db.entity_types.find_one_and_update(
                {"_id": et_id, "org_id": ctx_org},
                {"$inc": {"record_seq": 1}},
                projection={"key": 1, "record_seq": 1},
            )
            prefix = (seq.get("key") or "REC").upper()[:4]
            n = int(seq.get("record_seq") or 1)
            return f"{prefix}-{n:06d}"

        batch: list[dict] = []
        updates: list[tuple[str, dict]] = []
        BATCH = 200

        async def flush() -> None:
            nonlocal inserted, updated, errors
            if batch:
                try:
                    await db.records.insert_many(batch, ordered=False)
                    inserted += len(batch)
                except Exception:
                    # BulkWriteError or DuplicateKeyError — retry per-doc so
                    # we can count survivors and mark failures.
                    for d in batch:
                        try:
                            await db.records.insert_one(d)
                            inserted += 1
                        except Exception as exc:
                            errors += 1
                            error_lines.append([
                                "?", "_insert",
                                f"insert failed: {type(exc).__name__}",
                                d.get("record_number", "")[:32],
                            ])
                batch.clear()
            if updates:
                for rid, upd in updates:
                    try:
                        await db.records.update_one({"_id": rid, "org_id": ctx_org}, {"$set": upd})
                        updated += 1
                    except Exception as exc:
                        errors += 1
                        error_lines.append(["?", "_update",
                                            f"update failed: {type(exc).__name__}", rid])
                updates.clear()

        for row_idx, csv_row in enumerate(rows):
            processed += 1
            payload_fields: dict[str, Any] = {}
            title_val: Any = None
            tag_names: list[str] = []
            row_errors: list[dict] = []

            for header, val in csv_row.items():
                field_key = mapping.get(header)
                if not field_key or field_key in unsupported:
                    continue
                if field_key == "title":
                    title_val = val
                elif field_key == "tags":
                    tag_names = [x.strip() for x in str(val or "").split(",") if x.strip()]
                else:
                    payload_fields[field_key] = _coerce_import_value(defs_by_key.get(field_key), val)

            for fk, dv in default_values.items():
                if payload_fields.get(fk) in (None, "", []) and fk not in unsupported:
                    payload_fields[fk] = dv

            # Validate
            for fd in field_defs:
                if fd["key"] in unsupported:
                    continue
                v = payload_fields.get(fd["key"])
                if fd.get("required") and (v is None or v == ""):
                    row_errors.append({"field": fd["key"], "msg": "required"})
                    continue
                if v is None or v == "":
                    continue
                ok, coerced, err = _validate_field_value(fd, v)
                if not ok:
                    row_errors.append({"field": fd["key"], "msg": err or "invalid"})
                else:
                    payload_fields[fd["key"]] = coerced

            if row_errors:
                errors += 1
                for e in row_errors:
                    error_lines.append([str(row_idx), e["field"], e["msg"], str(csv_row.get(e["field"], ""))])
                continue

            # In-batch duplicate detection on `unique:true` fields.
            dup_key: str | None = None
            dup_first_row: int | None = None
            for uk in unique_field_keys:
                v = payload_fields.get(uk)
                if v is None or v == "":
                    continue
                first = seen_in_batch[uk].get(str(v))
                if first is not None:
                    dup_key = uk
                    dup_first_row = first
                    break
                seen_in_batch[uk][str(v)] = row_idx

            # Handle match / conflict
            existing_rid = None
            key_val = None
            if match_by:
                key_val = payload_fields.get(match_by) if match_by != "record_number" else csv_row.get(
                    next((h for h, k in mapping.items() if k == "record_number"), "")
                )
                if key_val and str(key_val) in existing_by_key:
                    existing_rid = existing_by_key[str(key_val)]["_id"]

            # In-batch duplicate resolution (mirrors plan-phase logic).
            if dup_key and not row_errors:
                if conflict_policy == "skip":
                    skipped += 1
                    continue
                if conflict_policy == "update" and match_by == dup_key:
                    # Collapse — target the DB row already scheduled or (if it
                    # was an insert) drop this duplicate.
                    tgt = existing_by_key.get(str(payload_fields.get(dup_key)), {}).get("_id")
                    if tgt and tgt not in collapsed_update_targets:
                        collapsed_update_targets.add(tgt)
                        # let the update path below run
                    else:
                        skipped += 1
                        continue
                else:
                    errors += 1
                    error_lines.append([
                        str(row_idx), dup_key,
                        f"duplicate unique value at row {row_idx + 1} "
                        f"(already at row {(dup_first_row or 0) + 1})",
                        str(payload_fields.get(dup_key, "")),
                    ])
                    continue

            # Tag lookup / auto-create
            tag_ids: list[str] = []
            for name in tag_names:
                tid = tag_map.get(name.lower())
                if tid:
                    tag_ids.append(tid)
                elif auto_create_tags:
                    new = {
                        "_id": str(uuid.uuid4()),
                        "org_id": ctx_org, "entity_type_id": et_id,
                        "name": name, "color": None, "usage_count": 0,
                        "created_at": now_iso(), "updated_at": now_iso(), "deleted_at": None,
                    }
                    try:
                        await db.tags.insert_one(new)
                        tag_map[name.lower()] = new["_id"]
                        tag_ids.append(new["_id"])
                    except Exception:
                        pass

            if existing_rid:
                if conflict_policy == "skip":
                    skipped += 1
                    continue
                if conflict_policy == "error":
                    errors += 1
                    error_lines.append([str(row_idx), match_by or "match", f"duplicate: {key_val}", ""])
                    continue
                # update
                upd_set = {
                    "fields": payload_fields,
                    "updated_at": now_iso(),
                }
                if title_val:
                    upd_set["title"] = str(title_val)
                if tag_ids:
                    upd_set["tag_ids"] = tag_ids
                updates.append((existing_rid, upd_set))
            else:
                rn = await _record_number(et_id)
                doc = {
                    "_id": str(uuid.uuid4()),
                    "org_id": ctx_org, "entity_type_id": et_id,
                    "title": str(title_val or "Untitled"),
                    "record_number": rn,
                    "fields": payload_fields,
                    "search_text": " ".join(str(v) for v in payload_fields.values() if v),
                    "category_ids": [], "tag_ids": tag_ids,
                    "version": 1,
                    "created_at": now_iso(), "updated_at": now_iso(), "deleted_at": None,
                }
                batch.append(doc)

            if len(batch) >= BATCH or len(updates) >= BATCH:
                await flush()
                # Update progress every batch
                await db.import_jobs.update_one(
                    {"_id": job_id},
                    {"$set": {"processed": processed, "inserted": inserted,
                              "updated": updated, "skipped": skipped, "errors": errors}},
                )
        await flush()

        # Persist error report if any
        error_key = None
        if len(error_lines) > 1:
            error_key = f"errors-{job_id}.csv"
            with open(_cache_path(error_key, ""), "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                for ln in error_lines:
                    w.writerow(ln)

        await db.import_jobs.update_one(
            {"_id": job_id},
            {"$set": {
                "status": "completed",
                "processed": processed, "inserted": inserted,
                "updated": updated, "skipped": skipped, "errors": errors,
                "error_report_key": error_key,
                "completed_at": now_iso(),
            }},
        )
        # Emit completion audit (background task with no BackgroundTasks — insert directly)
        try:
            await db.audit_logs.insert_one({
                "_id": str(uuid.uuid4()),
                "org_id": ctx_org,
                "actor_id": meta.get("user_id"),
                "action": "record.imported.completed",
                "target_type": "entity_type",
                "target_id": et_id,
                "diff": {"job_id": job_id, "inserted": inserted, "updated": updated,
                         "skipped": skipped, "errors": errors, "processed": processed},
                "ip": None, "ua": None,
                "ts": now_iso(),
            })
        except Exception:
            pass  # best-effort
    except Exception as exc:  # pragma: no cover
        await db.import_jobs.update_one(
            {"_id": job_id},
            {"$set": {"status": "failed", "completed_at": now_iso(),
                      "error_message": str(exc)[:500]}},
        )
    finally:
        _IMPORT_JOBS.pop(job_id, None)
